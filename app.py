import os
import time
import cv2
import pandas as pd
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
from ultralytics import YOLO
from datetime import datetime

# --- THƯ VIỆN TRANG TRÍ EXCEL ---
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'thadosoft_clone_key' 

# --- CẤU HÌNH ---
UPLOAD_FOLDER = 'static/uploads'
HISTORY_FOLDER = 'static/history'
MODEL_PATH = 'best.pt'

for folder in [UPLOAD_FOLDER, HISTORY_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Load Model
model = None
try:
    model = YOLO(MODEL_PATH)
    print("✅ Đã tải Model thành công!")
except:
    print("❌ Lỗi tải Model!")

# --- TRẠNG THÁI HỆ THỐNG ---
system_status = {
    "current_mode": 1,
    "trigger_capture": False,
    "last_update_time": 0,
    
    # Thêm trạng thái phần cứng để Web điều khiển
    "hardware": {
        "conveyor": False, # False = Tắt, True = Bật
        "valve": False     # False = Tắt, True = Bật
    },
    
    "stats": {
        "total": 0, "ok": 0, "ng": 0,
        "current_result": "WAITING", "ai_confidence": "---"
    }
}

# --- NHẬT KÝ HOẠT ĐỘNG ---
system_logs = []

def log_activity(username, action, detail=""):
    """Hàm ghi lại nhật ký hoạt động"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = {
        "time": now,
        "user": username,
        "action": action,
        "detail": detail
    }
    system_logs.insert(0, log_entry) 
    if len(system_logs) > 100:
        system_logs.pop()

# --- HÀM AI ---
def run_ai_inference(image_path):
    if not model: return "MODEL ERROR", "0%"
    results = model(image_path, conf=0.5)
    result = results[0]
    final_label = "NO DETECT"
    final_conf = 0.0

    if len(result.boxes) > 0:
        box = result.boxes[0]
        cls_id = int(box.cls[0])
        class_name = model.names[cls_id].lower()
        final_conf = float(box.conf[0])
        
        if 'ok' in class_name: final_label = "OK"
        elif 'ng' in class_name: final_label = "NG"
        else: final_label = class_name.upper()

        res_plotted = result.plot()
        cv2.imwrite(image_path, res_plotted)
    
    return final_label, f"{round(final_conf * 100, 1)}%"

# --- ROUTES ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == 'admin' and password == '123456':
            session['logged_in'] = True
            session['user'] = username
            log_activity(username, "Đăng nhập", "Truy cập hệ thống thành công")
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error="Sai tài khoản!")
    return render_template('login.html')

@app.route('/logout')
def logout():
    user = session.get('user', 'Unknown')
    log_activity(user, "Đăng xuất", "Thoát khỏi hệ thống")
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if not session.get('logged_in'): return redirect(url_for('login'))
    return render_template('index.html', user=session.get('user'))

# API LẤY NHẬT KÝ
@app.route('/get_logs')
def get_logs():
    return jsonify(system_logs)

# API LẤY DỮ LIỆU LỊCH SỬ CHI TIẾT
@app.route('/get_full_history')
def get_full_history():
    files = os.listdir(HISTORY_FOLDER)
    history_data = []
    for f in files:
        if f.endswith('.jpg') and f.startswith('img_'):
            try:
                parts = f.replace('.jpg', '').split('_')
                date_str = f"{parts[1][6:]}/{parts[1][4:6]}/{parts[1][:4]}"
                time_str = f"{parts[2][:2]}:{parts[2][2:4]}:{parts[2][4:]}"
                result = parts[3] if len(parts) > 3 else "Unknown"
                history_data.append({
                    "file": f, "date": date_str, "time": time_str, "result": result
                })
            except: continue
    history_data.sort(key=lambda x: (x['date'], x['time']), reverse=True)
    return jsonify(history_data)

# --- HÀM XUẤT EXCEL CHUYÊN NGHIỆP ---
@app.route('/export_excel')
def export_excel():
    user = session.get('user', 'Unknown')
    log_activity(user, "Xuất báo cáo", "Tải file Excel lịch sử")
    
    # 1. Thu thập dữ liệu
    files = os.listdir(HISTORY_FOLDER)
    data = []
    for f in files:
        if f.endswith('.jpg') and f.startswith('img_'):
            try:
                parts = f.replace('.jpg', '').split('_')
                date_str = f"{parts[1][6:]}/{parts[1][4:6]}/{parts[1][:4]}"
                time_str = f"{parts[2][:2]}:{parts[2][2:4]}:{parts[2][4:]}"
                result = parts[3] if len(parts) > 3 else "Unknown"
                
                data.append({
                    "STT": 0,
                    "Ngày": date_str,
                    "Giờ": time_str,
                    "Kết quả AI": result,
                    "Tên file ảnh": f
                })
            except: continue
    
    # Sắp xếp và đánh số STT
    data.sort(key=lambda x: (x['Ngày'], x['Giờ']), reverse=True)
    for i, item in enumerate(data):
        item["STT"] = i + 1
    
    df = pd.DataFrame(data)
    output_file = "Bao_Cao_San_Pham.xlsx"
    
    # 2. Xuất file với định dạng openpyxl
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Ghi dữ liệu từ dòng 7 (chừa chỗ cho Header)
        df.to_excel(writer, index=False, sheet_name='Báo Cáo', startrow=6)
        
        workbook = writer.book
        worksheet = writer.sheets['Báo Cáo']
        
        # --- TIÊU ĐỀ ---
        worksheet.merge_cells('A1:E2')
        cell_title = worksheet['A1']
        cell_title.value = "BÁO CÁO VẬN HÀNH HỆ THỐNG AI IOT"
        cell_title.font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
        cell_title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell_title.alignment = Alignment(horizontal='center', vertical='center')
        
        # --- THÔNG TIN META ---
        worksheet['A4'] = "Người xuất báo cáo:"
        worksheet['B4'] = user
        worksheet['A4'].font = Font(bold=True)
        
        worksheet['A5'] = "Thời gian xuất:"
        worksheet['B5'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        worksheet['A5'].font = Font(bold=True)
        
        # --- KẺ KHUNG VÀ CĂN CHỈNH ---
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col in worksheet.columns:
            column_letter = get_column_letter(col[0].column)
            max_length = 0
            for cell in col:
                if cell.row >= 7: # Kẻ bảng cho phần dữ liệu
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column_letter].width = max_length + 5

    return send_file(output_file, as_attachment=True)

# API GIAO TIẾP ESP32
@app.route('/upload', methods=['POST'])
def upload():
    file_data = request.data
    if file_data:
        # 1. Lưu ảnh
        live_path = os.path.join(UPLOAD_FOLDER, 'live.jpg')
        with open(live_path, 'wb') as f: f.write(file_data)
        
        # 2. Chạy AI để lấy kết quả (OK hoặc NG)
        label, conf = run_ai_inference(live_path)
        
        # 3. Cập nhật thống kê
        system_status["stats"]["total"] += 1
        system_status["stats"]["current_result"] = label
        system_status["stats"]["ai_confidence"] = conf
        if label == "OK": system_status["stats"]["ok"] += 1
        elif label == "NG": system_status["stats"]["ng"] += 1
        
        # 4. Cập nhật thời gian để Web hiển thị
        system_status["last_update_time"] = time.time() 

        # 5. Lưu lịch sử
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        history_path = os.path.join(HISTORY_FOLDER, f"img_{timestamp}_{label}.jpg")
        img_annotated = cv2.imread(live_path)
        if img_annotated is not None: cv2.imwrite(history_path, img_annotated)
            
        # --- QUAN TRỌNG: TRẢ VỀ KẾT QUẢ CHUẨN CHO ESP32 ---
        return label, 200 
        
    return "ERROR", 400

# API ĐIỀU KHIỂN PHẦN CỨNG (TEST TỪ WEB)
@app.route('/control_hardware', methods=['POST'])
def control_hardware():
    data = request.get_json()
    device = data.get('device') # 'conveyor' hoặc 'valve'
    action = data.get('action') # 'on' hoặc 'off'
    
    if device in system_status["hardware"]:
        system_status["hardware"][device] = (action == 'on')
        # Log lại để biết ai bấm
        user = session.get('user', 'Unknown')
        log_activity(user, "Điều khiển", f"Test {device} -> {action}")
        
    return jsonify({"status": "success", "state": system_status["hardware"]})

@app.route('/check_new_image')
def check_new_image():
    return jsonify({"last_update": system_status["last_update_time"], "stats": system_status["stats"]})

@app.route('/set_mode', methods=['POST'])
def set_mode():
    data = request.get_json()
    mode = int(data.get('mode', 1))
    system_status['current_mode'] = mode
    
    user = session.get('user', 'System')
    mode_name = "Tự động" if mode == 1 else "Stream" if mode == 2 else "Thủ công"
    log_activity(user, "Đổi chế độ", f"Chuyển sang chế độ {mode_name}")
    
    return jsonify({"status": "success"})

@app.route('/manual_capture', methods=['POST'])
def manual_capture():
    system_status['trigger_capture'] = True
    system_status["stats"]["current_result"] = "WAITING"
    
    user = session.get('user', 'Unknown')
    log_activity(user, "Thao tác thủ công", "Nhấn nút chụp ảnh")
    
    return jsonify({"status": "trigger_sent"})

# --- API CHECK STATUS (BAO GỒM CẢ LỆNH PHẦN CỨNG) ---
@app.route('/check_status', methods=['GET'])
def check_status():
    response = {
        "mode": system_status['current_mode'],
        "trigger": system_status['trigger_capture'],
        "hw": system_status["hardware"] # Gửi lệnh điều khiển phần cứng xuống ESP32
    }
    
    system_status['trigger_capture'] = False 
    
    # Tự động tắt lệnh kích Van sau khi gửi đi (để van không bị nóng)
    if system_status["hardware"]["valve"]:
         system_status["hardware"]["valve"] = False 

    return jsonify(response)

@app.route('/get_stats')
def get_stats(): return jsonify(system_status["stats"])

@app.route('/get_history')
def get_history():
    files = os.listdir(HISTORY_FOLDER)
    images = [f for f in files if f.endswith('.jpg')]
    images.sort(reverse=True)
    return jsonify(images[:12])

@app.route('/test_ai_upload', methods=['POST'])
def test_ai_upload():
    if 'file' not in request.files: return jsonify({"error": "No file"}), 400
    file = request.files['file']
    if file:
        live_path = os.path.join(UPLOAD_FOLDER, 'live.jpg')
        file.save(live_path)
        label, conf = run_ai_inference(live_path)
        system_status["last_update_time"] = time.time()
        system_status["stats"]["current_result"] = label
        system_status["stats"]["ai_confidence"] = conf
        
        user = session.get('user', 'Unknown')
        log_activity(user, "Test AI", f"Upload ảnh thủ công. KQ: {label}")
        
        return jsonify({"status": "success"})
    return jsonify({"error": "Error"}), 500

if __name__ == '__main__':
    # Lấy PORT từ biến môi trường của Render, nếu không có thì dùng 5000 (để chạy local vẫn được)
    port = int(os.environ.get("PORT", 5000))
    # Quan trọng: host phải là '0.0.0.0'
    app.run(host='0.0.0.0', port=port)

